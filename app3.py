import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go
import warnings
import base64
from pathlib import Path
from PIL import Image
warnings.filterwarnings("ignore")

from database import load_data

# ── AstraZeneca logo ──────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
AZ_LOGO  = Image.open(BASE_DIR / "assets" / "AZLogobyCopilot.png")

st.set_page_config(
    page_title="AZ · Tier Report",
    page_icon=AZ_LOGO,
    layout="wide",
)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Figtree:wght@300;400;600;700&family=Space+Mono:wght@400;700&display=swap');

html, body, [class*="css"] { font-family: 'Figtree', sans-serif; }

.stApp { background: #f4f6fb; color: #1a1a2e; }

/* ── Header ── */
.az-header {
    background: linear-gradient(135deg, #830051);
    border-radius: 12px;
    padding: 20px 28px;
    margin-bottom: 24px;
    display: flex;
    align-items: center;
    gap: 20px;
}
.az-header h1 {
    color: #ffffff;
    font-size: 1.5rem;
    font-weight: 700;
    margin: 0;
    letter-spacing: -0.01em;
}
.az-header p {
    color: rgba(255,255,255,0.65);
    font-size: 0.8rem;
    margin: 3px 0 0 0;
}

/* ── KPI Cards ── */
.kpi-card {
    background: #ffffff;
    border-radius: 10px;
    padding: 18px 16px;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    border-top: 4px solid #ccc;
    margin-bottom: 8px;
}
.kpi-card .kpi-num {
    font-family: 'Space Mono', monospace;
    font-size: 2.2rem;
    font-weight: 700;
    line-height: 1;
}
.kpi-card .kpi-label {
    font-size: 0.72rem;
    color: #666;
    margin-top: 5px;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    font-weight: 600;
}
.kpi-card .kpi-sub {
    font-size: 0.7rem;
    color: #999;
    margin-top: 2px;
}
.kpi-t1 { border-top-color: #830051; }
.kpi-t2 { border-top-color: #F0AB00; }
.kpi-t3 { border-top-color: #B0126D; }
.kpi-te { border-top-color: #C28F17; }
.num-t1 { color: #830051; }
.num-t2 { color: #F0AB00; }
.num-t3 { color: #B0126D; }
.num-te { color: #C28F17; }

/* ── Section headers ── */
.section-title {
    font-family: 'Space Mono', monospace;
    font-size: 0.68rem;
    color: #888;
    text-transform: uppercase;
    letter-spacing: 0.15em;
    margin-bottom: 10px;
    padding-bottom: 6px;
    border-bottom: 1px solid #e0e0e0;
}

/* ── Filter Pane ── */
[data-testid="stSidebar"] {
    background: #830051 !important;
}
[data-testid="stSidebar"] * { color: #ffffff !important; }
[data-testid="stSidebar"] .section-title {
    color: #F0AB00 !important;
    border-bottom-color: rgba(240,171,0,0.35) !important;
    font-weight: 700 !important;
}
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.15) !important; }

/* Sidebar input/select styling */
[data-testid="stSidebar"] [data-baseweb="select"] > div,
[data-testid="stSidebar"] [data-baseweb="input"] > div,
[data-testid="stSidebar"] textarea {
    background-color: rgba(255,255,255,0.12) !important;
    border-color: rgba(255,255,255,0.25) !important;
    color: #ffffff !important;
    border-radius: 6px !important;
}
[data-testid="stSidebar"] [data-baseweb="tag"] {
    background-color: #F0AB00 !important;
    color: #1a1a2e !important;
}
[data-testid="stSidebar"] label {
    color: rgba(255,255,255,0.85) !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
}

/* ── Buttons ── */
.stDownloadButton > button {
    background: #830051 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    width: 100% !important;
}
.stButton > button {
    background: #830051 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    width: 100% !important;
}
.stButton > button:hover,
.stDownloadButton > button:hover {
    background: #6a0040 !important;
    border-left: 3px solid #F0AB00 !important;
}

/* ── Tables ── */
div[data-testid="stDataFrame"] {
    border-radius: 10px;
    overflow: hidden;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
}

[data-testid="stDataFrame"] table {
    background-color: white !important;
    color: #1a1a1a !important;
}

[data-testid="stDataFrame"] thead tr th {
    background-color: #830051 !important;
    color: white !important;
    font-weight: 600 !important;
}

[data-testid="stDataFrame"] tbody tr td {
    background-color: white !important;
    color: #1a1a1a !important;
}

[data-testid="stDataFrame"] tbody tr:nth-child(even) td {
    background-color: #fdf4f9 !important;
}

[data-testid="stDataFrame"] tbody tr:hover td {
    background-color: #f5e6ef !important;
}
</style>
""", unsafe_allow_html=True)

# ── AZ logo base64 ────────────────────────────────────────────────────────────
def get_base64_image(image_path):
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode()

logo_b64 = get_base64_image(BASE_DIR / "assets" / "AZLogobyCopilot.png")

# ── TIER CONFIG ───────────────────────────────────────────────────────────────
TIER_THRESHOLDS = {"Tier 1": 0.51, "Tier 2": 0.81, "Tier 3": 0.96}
TIER_COLORS     = {
    "Tier 1":   "#830051",   # AZ primary
    "Tier 2":   "#F0AB00",   # AZ logo gold
    "Tier 3":   "#C21C78",   # AZ mid-pink
    "Tail-end": "#A97900",   # muted gold
}
MONTH_NAMES = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
               7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

# ── LOGIC ─────────────────────────────────────────────────────────────────────
def compute_tier(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["WeightedAmount"] = df["AmountUSD"] * df["ProductWeight"].fillna(1) if "ProductWeight" in df.columns else df["AmountUSD"]
    agg = df.groupby("AccountName")["WeightedAmount"].sum().reset_index().sort_values("WeightedAmount", ascending=False)
    total = agg["WeightedAmount"].sum()
    agg["% Grand Total"] = agg["WeightedAmount"] / total
    agg["Cumulative %"]  = agg["% Grand Total"].cumsum()

    def _tier(c):
        if c <= 0.51: return "Tier 1"
        if c <= 0.81: return "Tier 2"
        if c <= 0.96: return "Tier 3"
        return "Tail-end"

    agg["Tier"] = agg["Cumulative %"].apply(_tier)
    agg["Rank"] = range(1, len(agg) + 1)
    return agg[["Rank", "AccountName", "WeightedAmount", "% Grand Total", "Cumulative %", "Tier"]]


def build_pivot(df: pd.DataFrame, tier_df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["WeightedAmount"] = df["AmountUSD"] * df["ProductWeight"].fillna(1) if "ProductWeight" in df.columns else df["AmountUSD"]
    df = df.merge(tier_df[["AccountName", "Tier"]], on="AccountName", how="left")
    df["MonthLabel"] = df["ByYear"].astype(str) + "-" + df["ByMonth"].astype(str).str.zfill(2)

    pivot = (df.groupby(["Tier", "MonthLabel"])["WeightedAmount"].sum() / 1_000_000).reset_index()
    pivot_wide = pivot.pivot(index="Tier", columns="MonthLabel", values="WeightedAmount").fillna(0)
    pivot_wide = pivot_wide.reindex([t for t in ["Tier 1","Tier 2","Tier 3","Tail-end"] if t in pivot_wide.index])
    pivot_wide.columns.name = None
    pivot_wide.loc["Total"] = pivot_wide.sum()
    return pivot_wide.round(3)


def to_excel_bytes(pivot_df: pd.DataFrame) -> bytes:
    EXCEL_TIER_COLORS = {
        "Tier 1":   "FF830051",
        "Tier 2":   "FFF0AB00",
        "Tier 3":   "FFB0126D",
        "Tail-end": "FFC28F17",
        "Total":    "FF1A0050",
    }
    output = BytesIO()
    pivot_df.to_excel(output, sheet_name="Tier Pivot", index=True)
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    thin   = Side(style="thin", color="FFE0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor="FF830051")
        cell.font      = Font(bold=True, color="FFFFFFFF", name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    for row in ws.iter_rows(min_row=2):
        tier_name = row[0].value
        is_total  = tier_name == "Total"
        t_color   = EXCEL_TIER_COLORS.get(tier_name, "FFF4F6FB")
        for cell in row:
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column == 1:
                cell.fill = PatternFill("solid", fgColor=t_color)
                cell.font = Font(bold=True, color="FFFFFFFF", name="Calibri", size=10)
            else:
                cell.fill = PatternFill("solid", fgColor="FFFFFFFF" if not is_total else "FFF9F0F5")
                cell.font = Font(color="FF1A1A2E" if not is_total else "FF830051",
                                 bold=is_total, name="Calibri", size=10)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.000"

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4
    ws.row_dimensions[1].height = 22

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final.read()


def send_email(to_addr, subject, body, excel_bytes, filename):
    smtp_server = st.secrets["email"]["smtp_server"]
    smtp_port   = int(st.secrets["email"]["smtp_port"])
    sender      = st.secrets["email"]["sender"]
    password    = st.secrets["email"]["password"]
    msg = MIMEMultipart()
    msg["From"]    = sender
    msg["To"]      = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))
    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)
    with smtplib.SMTP(smtp_server, smtp_port) as s:
        s.ehlo(); s.starttls(); s.login(sender, password)
        s.sendmail(sender, to_addr, msg.as_string())


# ── HEADER ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="az-header">
    <div style="display:flex; align-items:center; gap:16px;">
        <img src="data:image/png;base64,{logo_b64}"
             style="height:56px; width:auto;" />
        <div>
            <h1>AstraZeneca · Tier Pivot Report</h1>
            <p>Sales performance by account tier &nbsp;·&nbsp; Monthly breakdown &nbsp;·&nbsp; USD millions</p>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── LOAD DATA ─────────────────────────────────────────────────────────────────
with st.spinner("🔄 Connecting to database..."):
    try:
        df_raw = load_data()
    except Exception as e:
        st.error(f"❌ Database connection failed: {e}")
        st.info("Running with sample data for preview.")
        np.random.seed(42)
        accounts = [f"Account {i:02d}" for i in range(1, 51)]
        rows = [{"CustomerID": a, "AccountName": a, "ByYear": 2025, "ByMonth": m,
                 "AmountUSD": np.random.exponential(50000),
                 "ProductWeight": np.random.uniform(0.8, 1.2),
                 "RegionID": np.random.choice(["North","South","Central"]),
                 "ChannelID": np.random.choice(["101","102","105"]),
                 "TerritoryID": np.random.choice(["T01","T02","T03"])}
                for a in accounts for m in range(1, 13)]
        df_raw = pd.DataFrame(rows)

# ── FILTER PANE ───────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<p class="section-title">⚙ Filters</p>', unsafe_allow_html=True)

    years = sorted(df_raw["ByYear"].dropna().unique().tolist(), reverse=True)
    sel_years = st.selectbox("📅 Year", years, index=0)

    sel_months = st.multiselect(
        "🗓 Month", list(range(1, 13)),
        default=list(range(1, 13)),
        format_func=lambda x: MONTH_NAMES[x]
    )

    regions = sorted(df_raw["RegionID"].dropna().unique().tolist())
    sel_regions = st.multiselect("🗺 Region", regions, default=regions)

    channels = sorted(df_raw["ChannelID"].dropna().unique().tolist())
    sel_channels = st.multiselect("🏥 Channel", channels, default=channels)

    territories = sorted(df_raw["TerritoryID"].dropna().unique().tolist())
    sel_territories = st.multiselect("📍 Territory", territories, default=territories)

    st.markdown("---")
    st.markdown('<p class="section-title">✉ Email Export</p>', unsafe_allow_html=True)
    email_to      = st.text_input("Recipient")
    email_subject = st.text_input("Subject", value="AZ Tier Pivot Report")
    email_body    = st.text_area("Message", value="Please find the Tier Pivot Report attached.", height=80)

# ── FILTER ────────────────────────────────────────────────────────────────────
df = df_raw.copy()
df = df[df["ByYear"] == sel_years]
if sel_months:      df = df[df["ByMonth"].isin(sel_months)]
if sel_regions:     df = df[df["RegionID"].isin(sel_regions)]
if sel_channels:    df = df[df["ChannelID"].isin(sel_channels)]
if sel_territories: df = df[df["TerritoryID"].isin(sel_territories)]

if df.empty:
    st.warning("⚠️ No data matches the selected filters.")
    st.stop()

# ── COMPUTE ───────────────────────────────────────────────────────────────────
tier_df  = compute_tier(df)
pivot_df = build_pivot(df, tier_df)

# ── KPI CARDS ─────────────────────────────────────────────────────────────────
counts  = tier_df["Tier"].value_counts()
revenue = tier_df.groupby("Tier")["WeightedAmount"].sum() / 1_000_000

c1, c2, c3, c4 = st.columns(4)
for col, tier, cls, ncls in zip(
    [c1, c2, c3, c4],
    ["Tier 1", "Tier 2", "Tier 3", "Tail-end"],
    ["kpi-t1","kpi-t2","kpi-t3","kpi-te"],
    ["num-t1","num-t2","num-t3","num-te"]
):
    with col:
        rev = revenue.get(tier, 0)
        st.markdown(f"""
        <div class="kpi-card {cls}">
            <div class="kpi-num {ncls}">{counts.get(tier, 0)}</div>
            <div class="kpi-label">{tier}</div>
            <div class="kpi-sub">USD {rev:.2f}M</div>
        </div>
        """, unsafe_allow_html=True)

st.markdown("---")

# ── TIER RANKING TABLE ─────────────────────────────────────────────────────────
st.markdown('<p class="section-title">📋 Tier Ranking — Account List</p>', unsafe_allow_html=True)

display_tier = tier_df.copy()
display_tier["WeightedAmount"] = (display_tier["WeightedAmount"] / 1_000_000).round(3)
display_tier["% Grand Total"]  = (display_tier["% Grand Total"] * 100).round(2).astype(str) + "%"
display_tier["Cumulative %"]   = (display_tier["Cumulative %"] * 100).round(1).astype(str) + "%"
display_tier = display_tier.rename(columns={"WeightedAmount": "Revenue (USD M)"})

st.dataframe(
    display_tier,
    use_container_width=True,
    height=350,
    hide_index=True,
    column_config={
        "Rank":            st.column_config.NumberColumn("Rank", width="small"),
        "AccountName":     st.column_config.TextColumn("Account Name", width="large"),
        "Revenue (USD M)": st.column_config.NumberColumn("Revenue (USD M)", format="%.3f"),
        "Tier":            st.column_config.TextColumn("Tier", width="small"),
    }
)

st.markdown("---")

# ── PIVOT TABLE ───────────────────────────────────────────────────────────────
st.markdown('<p class="section-title">📊 Pivot: Tier × Month (USD Millions)</p>', unsafe_allow_html=True)
st.dataframe(
    pivot_df.style.format("{:.3f}")
        .background_gradient(cmap="RdPu", axis=1, subset=pd.IndexSlice[:"Tail-end", :]),
    use_container_width=True,
    height=230,
)

st.markdown("---")

# ── CHARTS ────────────────────────────────────────────────────────────────────
st.markdown('<p class="section-title">📈 Charts</p>', unsafe_allow_html=True)

# Shared layout defaults
_axis_style = dict(
    title_font=dict(color="#1a1a2e", family="Figtree"),
    tickfont=dict(color="#1a1a2e", family="Figtree"),
    gridcolor="#f0e6ec",
    linecolor="#e0c8d5",
)
_layout_base = dict(
    plot_bgcolor="white",
    paper_bgcolor="white",
    font=dict(family="Figtree", color="#1a1a2e"),
)

chart_col1, chart_col2 = st.columns(2)

# Chart 1 — Monthly Revenue by Tier (line)
with chart_col1:
    st.markdown("**Monthly Revenue by Tier (USD M)**")
    pivot_no_total = pivot_df.drop(index="Total", errors="ignore")
    melted = pivot_no_total.reset_index().melt(id_vars="Tier", var_name="Month", value_name="Revenue")
    fig1 = px.line(
        melted, x="Month", y="Revenue", color="Tier",
        color_discrete_map=TIER_COLORS,
        markers=True,
        template="plotly_white",
    )
    fig1.update_layout(
        **_layout_base,
        margin=dict(l=10, r=10, t=10, b=40),
        legend=dict(
            orientation="h", yanchor="bottom", y=1.02,
            font=dict(color="#1a1a2e", family="Figtree"),
            title=dict(font=dict(color="#1a1a2e")),
        ),
        xaxis=dict(tickangle=350, **_axis_style),
        yaxis=dict(**_axis_style),
        height=320,
    )
    st.plotly_chart(fig1, use_container_width=True)

# Chart 2 — Revenue Share by Tier (donut)
with chart_col2:
    st.markdown("**Revenue Share by Tier**")
    share_data = tier_df.groupby("Tier")["WeightedAmount"].sum().reset_index()
    fig2 = px.pie(
        share_data, values="WeightedAmount", names="Tier",
        color="Tier", color_discrete_map=TIER_COLORS,
        hole=0.55,
        template="plotly_white",
    )
    fig2.update_traces(
        textposition="outside",
        textinfo="percent+label",
        textfont=dict(family="Figtree", color="#1a1a2e"),
    )
    fig2.update_layout(
        **_layout_base,
        margin=dict(l=10, r=10, t=10, b=10),
        showlegend=False,
        height=320,
    )
    st.plotly_chart(fig2, use_container_width=True)

chart_col3, chart_col4 = st.columns(2)

# Chart 3 — Account Count by Tier (bar)
with chart_col3:
    st.markdown("**Account Count by Tier**")
    count_data = tier_df["Tier"].value_counts().reset_index()
    count_data.columns = ["Tier", "Count"]
    fig3 = px.bar(
        count_data, x="Tier", y="Count",
        color="Tier", color_discrete_map=TIER_COLORS,
        text="Count",
        template="plotly_white",
    )
    fig3.update_traces(
        textposition="outside",
        textfont=dict(family="Figtree", color="#1a1a2e"),
    )
    fig3.update_layout(
        **_layout_base,
        showlegend=False,
        margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(**_axis_style),
        yaxis=dict(**_axis_style),
        height=300,
    )
    st.plotly_chart(fig3, use_container_width=True)

# Chart 4 — Top 20 Accounts Cumulative Revenue %
with chart_col4:
    st.markdown("**Top 20 Accounts — Cumulative Revenue %**")
    top20 = tier_df.head(20).copy()
    top20["CumPct"] = top20["Cumulative %"] * 100 if top20["Cumulative %"].max() <= 1 else top20["Cumulative %"]
    pct_col = top20["% Grand Total"] * 100 if top20["% Grand Total"].max() <= 1 else top20["% Grand Total"]

    fig4 = go.Figure()
    fig4.add_trace(go.Bar(
        x=top20["AccountName"].str[:20],
        y=pct_col,
        marker_color=[TIER_COLORS.get(t, "#999") for t in top20["Tier"]],
        name="% Revenue",
    ))
    fig4.update_layout(
        **_layout_base,
        xaxis=dict(tickangle=45, tickfont=dict(color="#1a1a2e", size=9, family="Figtree"),
                   linecolor="#e0c8d5", gridcolor="#f0e6ec"),
        yaxis=dict(title="% Revenue", **_axis_style),
        margin=dict(l=10, r=10, t=10, b=80),
        showlegend=False,
        height=300,
    )
    st.plotly_chart(fig4, use_container_width=True)

st.markdown("---")

# ── EXPORT ────────────────────────────────────────────────────────────────────
excel_bytes = to_excel_bytes(pivot_df)
filename    = f"AZ_TierPivot_{sel_years}.xlsx"

col_dl, col_email = st.columns(2)
with col_dl:
    st.download_button(
        "⬇ Download Excel",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with col_email:
    if st.button("✉ Send Email", use_container_width=True):
        if not email_to:
            st.error("Please enter a recipient email.")
        else:
            with st.spinner("Sending..."):
                try:
                    send_email(email_to, email_subject, email_body, excel_bytes, filename)
                    st.success(f"✅ Sent to {email_to}")
                except Exception as e:
                    st.error(f"❌ Failed: {e}")


