import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Tier Pivot Report",
    page_icon="📊",
    layout="wide",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

.stApp {
    background: #0f0f14;
    color: #e8e8f0;
}

h1, h2, h3 {
    font-family: 'IBM Plex Mono', monospace;
    letter-spacing: -0.02em;
}

.header-block {
    border-left: 3px solid #c8a951;
    padding: 8px 16px;
    margin-bottom: 24px;
    background: #16161e;
}

.header-block h1 {
    color: #c8a951;
    font-size: 1.6rem;
    margin: 0;
}

.header-block p {
    color: #888;
    font-size: 0.82rem;
    margin: 4px 0 0 0;
    font-family: 'IBM Plex Mono', monospace;
}

.tier-card {
    background: #16161e;
    border: 1px solid #2a2a38;
    border-radius: 4px;
    padding: 16px;
    text-align: center;
    margin-bottom: 8px;
}

.tier-card .count {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 2rem;
    font-weight: 600;
    line-height: 1;
}

.tier-card .label {
    font-size: 0.75rem;
    color: #888;
    margin-top: 4px;
    text-transform: uppercase;
    letter-spacing: 0.1em;
}

.t1 { color: #4caf8a; border-top: 2px solid #4caf8a; }
.t2 { color: #c8a951; border-top: 2px solid #c8a951; }
.t3 { color: #5b9bd5; border-top: 2px solid #5b9bd5; }
.te { color: #e05c5c; border-top: 2px solid #e05c5c; }

.stSelectbox > div, .stMultiSelect > div {
    background: #16161e !important;
    border-color: #2a2a38 !important;
}

.stButton > button {
    background: #c8a951;
    color: #0f0f14;
    border: none;
    font-family: 'IBM Plex Mono', monospace;
    font-weight: 600;
    letter-spacing: 0.05em;
    border-radius: 3px;
    padding: 10px 24px;
    width: 100%;
}

.stButton > button:hover {
    background: #e0c06a;
}

div[data-testid="stDataFrame"] {
    border: 1px solid #2a2a38;
    border-radius: 4px;
}

.section-label {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.7rem;
    color: #888;
    text-transform: uppercase;
    letter-spacing: 0.15em;
    margin-bottom: 8px;
}

.divider {
    border: none;
    border-top: 1px solid #2a2a38;
    margin: 24px 0;
}
</style>
""", unsafe_allow_html=True)


# ── DB Connection ─────────────────────────────────────────────────────────────
@st.cache_resource
def get_connection():
    server   = st.secrets["db"]["server"]
    database = st.secrets["db"]["database"]
    username = st.secrets["db"]["username"]
    password = st.secrets["db"]["password"]
    conn_str = (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        f"SERVER={server};DATABASE={database};UID={username};PWD={password};"
    )
    return pyodbc.connect(conn_str)


@st.cache_data(ttl=3600)
def load_data():
    conn = get_connection()

    df_customer = pd.read_sql("""
        SELECT T0.CustomerID, T0.CustomerName, T0.Address, T0.RegionID,
               T0.CityID, T0.AccountID, T1.AccountName,
               CASE
                   WHEN T0.ChannelName = '21-Gov. Hosp - Dept'    THEN '101'
                   WHEN T0.ChannelName = '22-Gov. Hosp - Phar'    THEN '102'
                   WHEN T0.ChannelName = '23-Private Hosp - Dept' THEN '104'
                   WHEN T0.ChannelName = '24-Private Hosp - Phar' THEN '106'
                   WHEN T0.ChannelName = '25-Polyclinic'          THEN '103'
                   WHEN T0.ChannelName = '26-Pharmacy'            THEN '105'
                   WHEN T0.ChannelName = '27-Wholesalers'         THEN '107'
                   WHEN T0.ChannelName = '28-Monoclinic'          THEN '108'
                   WHEN T0.ChannelName = '29-Pharmacy Chain'      THEN '109'
               END AS ChannelID
        FROM [VNDB_O365].[dbo].[vw_Sipro_Customer] AS T0
        LEFT JOIN [VNDB_O365].[dbo].[vw_ParentAccount] T1
            ON T0.CUSTOMERID = T1.CUSTOMERID
    """, conn)

    df_sales = pd.read_sql("""
        SELECT T0.CustomerID, T0.Qty, T1.TerritoryID, T0.Cust_Brand,
               T0.BrandID, T0.ProductID, T0.ByMonth, T0.ByYear,
               T0.YYMM, T0.AmountUSD, T1.ProductWeight
        FROM (
            SELECT CustomerID, AmountUSD, Qty, BrandID,
                   ProductID, ByMonth, ByYear,
                   YYMM, CONCAT(CustomerID, Brand) AS Cust_Brand
            FROM [VNDB_O365].[dbo].[vw_SIpro_CustomerProductByCutoff]
        ) AS T0
        LEFT JOIN (
            SELECT TeamID, TerritoryID, CustomerID, BrandID,
                   ProductWeight, Brand,
                   CONCAT(CustomerID, Brand) AS Cust_Brand
            FROM [VNDB_O365].[dbo].[vw_SIpro_TerritoryAlignment]
        ) AS T1 ON T0.Cust_Brand = T1.Cust_Brand
    """, conn)

    df = df_sales.merge(df_customer, on="CustomerID", how="left")
    return df


# ── Tier Logic ────────────────────────────────────────────────────────────────
def compute_tier(df_filtered):
    """Compute Tier ranking from filtered dataframe."""
    # Weighted amount = AmountUSD * ProductWeight (fallback to AmountUSD)
    if "ProductWeight" in df_filtered.columns:
        df_filtered = df_filtered.copy()
        df_filtered["WeightedAmount"] = (
            df_filtered["AmountUSD"] * df_filtered["ProductWeight"].fillna(1)
        )
    else:
        df_filtered["WeightedAmount"] = df_filtered["AmountUSD"]

    # Aggregate by AccountName
    acc_agg = (
        df_filtered.groupby("AccountName")["WeightedAmount"]
        .sum()
        .reset_index()
        .sort_values("WeightedAmount", ascending=False)
    )
    total = acc_agg["WeightedAmount"].sum()
    acc_agg["PctGrandTotal"] = acc_agg["WeightedAmount"] / total
    acc_agg["CumulativeSum"]  = acc_agg["PctGrandTotal"].cumsum()

    def assign_tier(c):
        if   c <= 0.51: return "Tier 1"
        elif c <= 0.81: return "Tier 2"
        elif c <= 0.96: return "Tier 3"
        else:           return "Tail-end"

    acc_agg["Tier"] = acc_agg["CumulativeSum"].apply(assign_tier)
    return acc_agg[["AccountName", "Tier", "PctGrandTotal", "CumulativeSum"]]


# ── Pivot Builder ─────────────────────────────────────────────────────────────
def build_pivot(df_filtered, tier_df):
    """Build Tier × Month pivot (AmountUSD / 1,000,000)."""
    df = df_filtered.merge(tier_df[["AccountName", "Tier"]], on="AccountName", how="left")

    if "ProductWeight" in df.columns:
        df["WeightedAmount"] = df["AmountUSD"] * df["ProductWeight"].fillna(1)
    else:
        df["WeightedAmount"] = df["AmountUSD"]

    df["MonthLabel"] = df["ByYear"].astype(str) + "-" + df["ByMonth"].astype(str).str.zfill(2)

    pivot = (
        df.groupby(["Tier", "MonthLabel"])["WeightedAmount"]
        .sum()
        .reset_index()
    )
    pivot["WeightedAmount"] = pivot["WeightedAmount"] / 1_000_000

    pivot_wide = pivot.pivot(index="Tier", columns="MonthLabel", values="WeightedAmount").fillna(0)

    # Order tiers
    tier_order = ["Tier 1", "Tier 2", "Tier 3", "Tail-end"]
    pivot_wide = pivot_wide.reindex([t for t in tier_order if t in pivot_wide.index])
    pivot_wide.columns.name = None

    # Add Total row
    pivot_wide.loc["Total"] = pivot_wide.sum()
    return pivot_wide.round(3)


# ── Excel Export ──────────────────────────────────────────────────────────────
TIER_COLORS = {
    "Tier 1":   "FF4CAF8A",
    "Tier 2":   "FFC8A951",
    "Tier 3":   "FF5B9BD5",
    "Tail-end": "FFE05C5C",
    "Total":    "FF2A2A38",
}

def to_excel_bytes(pivot_df: pd.DataFrame) -> bytes:
    output = BytesIO()
    pivot_df.to_excel(output, sheet_name="Tier Pivot", index=True)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="FF0F0F14")
    header_font = Font(bold=True, color="FFC8A951", name="Calibri")
    thin = Side(style="thin", color="FF2A2A38")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Style header row
    for cell in ws[1]:
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Style data rows
    for row in ws.iter_rows(min_row=2):
        tier_name = row[0].value
        color     = TIER_COLORS.get(tier_name, "FF16161E")
        row_fill  = PatternFill("solid", fgColor=color)
        is_total  = tier_name == "Total"

        for cell in row:
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column == 1:
                cell.fill = row_fill
                cell.font = Font(bold=True, color="FF0F0F14" if not is_total else "FFE8E8F0", name="Calibri")
            else:
                cell.fill = PatternFill("solid", fgColor="FF16161E")
                cell.font = Font(color="FFE8E8F0" if not is_total else "FFC8A951",
                                 bold=is_total, name="Calibri")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.000"

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    ws.row_dimensions[1].height = 22

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final.read()


# ── Email Sender ──────────────────────────────────────────────────────────────
def send_email_outlook(to_addr: str, subject: str, body: str, excel_bytes: bytes, filename: str):
    smtp_server = st.secrets["email"]["smtp_server"]   # smtp.office365.com
    smtp_port   = int(st.secrets["email"]["smtp_port"]) # 587
    sender      = st.secrets["email"]["sender"]
    password    = st.secrets["email"]["password"]

    msg = MIMEMultipart()
    msg["From"]    = sender
    msg["To"]      = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.ehlo()
        server.starttls()
        server.login(sender, password)
        server.sendmail(sender, to_addr, msg.as_string())


# ── UI ────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-block">
    <h1>📊 TIER PIVOT REPORT</h1>
    <p>Sales performance by account tier · Monthly breakdown </p>
</div>
""", unsafe_allow_html=True)

# Load data
with st.spinner("Loading data from database..."):
    try:
        df_raw = load_data()
        data_ok = True
    except Exception as e:
        st.error(f"❌ Database connection failed: {e}")
        st.info("Running with sample data for preview.")
        data_ok = False
        # Sample data fallback
        np.random.seed(42)
        accounts = [f"Account {i}" for i in range(1, 51)]
        months   = [(2025, m) for m in range(1, 13)]
        rows = []
        for acc in accounts:
            for yr, mo in months:
                rows.append({
                    "CustomerID": acc, "AccountName": acc,
                    "ByYear": yr, "ByMonth": mo,
                    "AmountUSD": np.random.exponential(50000),
                    "ProductWeight": np.random.uniform(0.8, 1.2),
                    "RegionID": np.random.choice(["North", "South", "Central"]),
                    "ChannelID": np.random.choice(["101", "102", "105"]),
                    "TerritoryID": np.random.choice(["T01", "T02", "T03"]),
                })
        df_raw = pd.DataFrame(rows)

# ── Sidebar Filters ───────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<p class="section-label">⚙ Filters</p>', unsafe_allow_html=True)

    years = sorted(df_raw["ByYear"].dropna().unique().tolist(), reverse=True)
    sel_years = st.multiselect("Year", years, default=years[:1])

    months_all = list(range(1, 13))
    month_names = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                   7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
    sel_months = st.multiselect(
        "Month", months_all,
        default=months_all,
        format_func=lambda x: month_names[x]
    )

    regions = sorted(df_raw["RegionID"].dropna().unique().tolist())
    sel_regions = st.multiselect("Region", regions, default=regions)

    channels = sorted(df_raw["ChannelID"].dropna().unique().tolist())
    sel_channels = st.multiselect("Channel", channels, default=channels)

    territories = sorted(df_raw["TerritoryID"].dropna().unique().tolist())
    sel_territories = st.multiselect("Territory", territories, default=territories)

    st.markdown('<hr class="divider">', unsafe_allow_html=True)
    st.markdown('<p class="section-label">✉ Email Export</p>', unsafe_allow_html=True)
    email_to      = st.text_input("Recipient email")
    email_subject = st.text_input("Subject", value="Tier Pivot Report")
    email_body    = st.text_area("Message", value="Please find the Tier Pivot Report attached.", height=80)

# ── Apply Filters ─────────────────────────────────────────────────────────────
df_filtered = df_raw.copy()
if sel_years:       df_filtered = df_filtered[df_filtered["ByYear"].isin(sel_years)]
if sel_months:      df_filtered = df_filtered[df_filtered["ByMonth"].isin(sel_months)]
if sel_regions:     df_filtered = df_filtered[df_filtered["RegionID"].isin(sel_regions)]
if sel_channels:    df_filtered = df_filtered[df_filtered["ChannelID"].isin(sel_channels)]
if sel_territories: df_filtered = df_filtered[df_filtered["TerritoryID"].isin(sel_territories)]

# ── Compute Tier & Pivot ──────────────────────────────────────────────────────
if df_filtered.empty:
    st.warning("No data matches the selected filters.")
    st.stop()

tier_df  = compute_tier(df_filtered)
pivot_df = build_pivot(df_filtered, tier_df)

# ── Tier Summary Cards ────────────────────────────────────────────────────────
counts = tier_df["Tier"].value_counts()
c1, c2, c3, c4 = st.columns(4)
for col, tier, cls in zip(
    [c1, c2, c3, c4],
    ["Tier 1", "Tier 2", "Tier 3", "Tail-end"],
    ["t1", "t2", "t3", "te"]
):
    with col:
        st.markdown(f"""
        <div class="tier-card {cls}">
            <div class="count">{counts.get(tier, 0)}</div>
            <div class="label">{tier}</div>
        </div>
        """, unsafe_allow_html=True)

st.markdown('<hr class="divider">', unsafe_allow_html=True)

# ── Tier Ranking Table ────────────────────────────────────────────────────────
col_left, col_right = st.columns([1, 2])

with col_left:
    st.markdown('<p class="section-label">Tier Ranking</p>', unsafe_allow_html=True)
    display_tier = tier_df.copy()
    display_tier["PctGrandTotal"] = (display_tier["PctGrandTotal"] * 100).round(2).astype(str) + "%"
    display_tier["CumulativeSum"] = (display_tier["CumulativeSum"] * 100).round(1).astype(str) + "%"
    st.dataframe(display_tier, use_container_width=True, height=400, hide_index=True)

with col_right:
    st.markdown('<p class="section-label">Pivot: Tier (Monhtly Sales)</p>', unsafe_allow_html=True)
    st.dataframe(
        pivot_df.style.format("{:.3f}").highlight_max(axis=1, color="#1e2a1e"),
        use_container_width=True,
        height=400
    )

st.markdown('<hr class="divider">', unsafe_allow_html=True)

# ── Export & Email ────────────────────────────────────────────────────────────
excel_bytes = to_excel_bytes(pivot_df)
filename    = f"TierPivot_{'_'.join(map(str, sel_years))}.xlsx"

col_dl, col_email = st.columns(2)

with col_dl:
    st.download_button(
        label="⬇ Download Excel",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

with col_email:
    if st.button("✉ Send Email", use_container_width=True):
        if not email_to:
            st.error("Please enter a recipient email address.")
        else:
            with st.spinner("Sending email..."):
                try:
                    send_email_outlook(email_to, email_subject, email_body, excel_bytes, filename)
                    st.success(f"✅ Email sent to {email_to}")
                except Exception as e:
                    st.error(f"❌ Failed to send email: {e}")
